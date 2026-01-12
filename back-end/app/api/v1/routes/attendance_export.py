"""
Attendance Export API Route
Exports attendance records for a study as Excel file.
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.db.odm.volunteer_attendance import VolunteerAttendance
from app.api.v1 import deps

router = APIRouter()


@router.get("/export/{study_code}")
async def export_attendance(
    study_code: str,
    user: dict = Depends(deps.get_current_user)
):
    """
    Export attendance records for a study as Excel file.
    
    Pulls data from:
    1. VolunteerAttendance collection (for volunteers with check-in/out records)
    2. AssignedStudy collection (for all assigned volunteers)
    
    This ensures ALL volunteers assigned to a study appear in the report,
    even if they haven't checked in yet.
    """
    from app.db.odm.assigned_study import AssignedStudy
    
    # Fetch attendance records (volunteers who have checked in/out)
    attendance_records = await VolunteerAttendance.find(
        VolunteerAttendance.study_code == study_code
    ).to_list()
    
    # Fetch assigned study records (all volunteers assigned to this study)
    assigned_volunteers = await AssignedStudy.find(
        AssignedStudy.study_code == study_code
    ).to_list()
    
    if not attendance_records and not assigned_volunteers:
        raise HTTPException(
            status_code=404,
            detail=f"No volunteers found for study: {study_code}"
        )
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Define headers
    headers = [
        "Study Code",
        "Study Name", 
        "Volunteer ID",
        "Volunteer Name",
        "Date",
        "Check-In Time",
        "Check-Out Time",
        "Duration (Hours)",
        "Status"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Study Code
    ws.column_dimensions['B'].width = 30  # Study Name
    ws.column_dimensions['C'].width = 15  # Volunteer ID
    ws.column_dimensions['D'].width = 25  # Volunteer Name
    ws.column_dimensions['E'].width = 12  # Date
    ws.column_dimensions['F'].width = 18  # Check-In Time
    ws.column_dimensions['G'].width = 18  # Check-Out Time
    ws.column_dimensions['H'].width = 15  # Duration
    ws.column_dimensions['I'].width = 15  # Status
    
    # Get study name from first available record
    study_name = study_code  # Default fallback
    if attendance_records:
        study_name = attendance_records[0].study_name
    elif assigned_volunteers:
        study_name = assigned_volunteers[0].study_name
    
    # Track which volunteers we've added (by volunteer_id)
    processed_volunteers = set()
    row_num = 2
    
    # First, add all attendance records (volunteers who have checked in/out)
    for volunteer_record in attendance_records:
        processed_volunteers.add(volunteer_record.volunteer_id)
        
        # Process all attendance logs for this volunteer
        if volunteer_record.attendance_logs and len(volunteer_record.attendance_logs) > 0:
            for log in volunteer_record.attendance_logs:
                check_in = log.get("check_in")
                check_out = log.get("check_out")
                duration = log.get("duration_hours", 0)
                
                # Format timestamps
                date_str = check_in.strftime("%Y-%m-%d") if check_in else "-"
                check_in_str = check_in.strftime("%Y-%m-%d %H:%M:%S") if check_in else "-"
                check_out_str = check_out.strftime("%Y-%m-%d %H:%M:%S") if check_out else "-"
                
                # Write row
                ws.cell(row=row_num, column=1).value = study_code
                ws.cell(row=row_num, column=2).value = study_name
                ws.cell(row=row_num, column=3).value = volunteer_record.volunteer_id
                ws.cell(row=row_num, column=4).value = volunteer_record.volunteer_name
                ws.cell(row=row_num, column=5).value = date_str
                ws.cell(row=row_num, column=6).value = check_in_str
                ws.cell(row=row_num, column=7).value = check_out_str
                ws.cell(row=row_num, column=8).value = duration
                ws.cell(row=row_num, column=9).value = "Completed"
                
                # Center align date, duration, and status columns
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")
                
                row_num += 1
        else:
            # If volunteer has no logs yet but is currently active
            if volunteer_record.is_active and volunteer_record.check_in_time:
                check_in = volunteer_record.check_in_time
                date_str = check_in.strftime("%Y-%m-%d")
                check_in_str = check_in.strftime("%Y-%m-%d %H:%M:%S")
                
                ws.cell(row=row_num, column=1).value = study_code
                ws.cell(row=row_num, column=2).value = study_name
                ws.cell(row=row_num, column=3).value = volunteer_record.volunteer_id
                ws.cell(row=row_num, column=4).value = volunteer_record.volunteer_name
                ws.cell(row=row_num, column=5).value = date_str
                ws.cell(row=row_num, column=6).value = check_in_str
                ws.cell(row=row_num, column=7).value = "-"
                ws.cell(row=row_num, column=8).value = "-"
                ws.cell(row=row_num, column=9).value = "Currently Checked In"
                
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")
                
                row_num += 1
    
    # Second, add assigned volunteers who haven't checked in yet
    for assigned in assigned_volunteers:
        if assigned.volunteer_id not in processed_volunteers:
            # This volunteer is assigned but has never checked in
            ws.cell(row=row_num, column=1).value = study_code
            ws.cell(row=row_num, column=2).value = study_name
            ws.cell(row=row_num, column=3).value = assigned.volunteer_id
            ws.cell(row=row_num, column=4).value = assigned.volunteer_name
            ws.cell(row=row_num, column=5).value = "-"
            ws.cell(row=row_num, column=6).value = "-"
            ws.cell(row=row_num, column=7).value = "-"
            ws.cell(row=row_num, column=8).value = "0"
            ws.cell(row=row_num, column=9).value = "Assigned (No Attendance)"
            
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")
            
            # Gray out rows for volunteers with no attendance
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).font = Font(color="999999", italic=True)
            
            row_num += 1
    
    # Save workbook to bytes buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename with timestamp
    filename = f"Attendance_{study_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as downloadable file
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
